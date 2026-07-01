import io
import logging
import re
import uuid
from datetime import datetime
from typing import Optional

import asyncpg
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from pydantic import BaseModel

from auth import require_admin
from db import get_pool

log = logging.getLogger(__name__)
router = APIRouter()

# ── Schema (auto-created on first request) ────────────────────────────────────

_DDL = """
CREATE TABLE IF NOT EXISTS fbdi_templates (
    id                UUID        PRIMARY KEY DEFAULT gen_random_uuid(),
    name              TEXT        NOT NULL,
    original_filename TEXT        NOT NULL,
    sheet_count       INTEGER     NOT NULL DEFAULT 0,
    column_count      INTEGER     NOT NULL DEFAULT 0,
    status            TEXT        NOT NULL DEFAULT 'active',
    created_at        TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    updated_at        TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS fbdi_columns (
    id          UUID        PRIMARY KEY DEFAULT gen_random_uuid(),
    template_id UUID        NOT NULL REFERENCES fbdi_templates(id) ON DELETE CASCADE,
    sheet_name  TEXT        NOT NULL,
    column_name TEXT        NOT NULL,
    position    INTEGER     NOT NULL,
    notes       TEXT,
    created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_fbdi_cols_tpl ON fbdi_columns(template_id);
"""

# ── Parser ────────────────────────────────────────────────────────────────────

_ORACLE_RE = re.compile(r'^[A-Z][A-Z0-9_]{1,59}$')


def _is_oracle_row(row: tuple) -> bool:
    """True if >50 % of non-null cells look like Oracle column names."""
    non_null = [str(v).strip() for v in row if v is not None and str(v).strip()]
    if len(non_null) < 2:
        return False
    return sum(_ORACLE_RE.match(v) is not None for v in non_null) / len(non_null) > 0.5


def _parse_workbook(file_bytes: bytes) -> list[dict]:
    """
    Parse an FBDI Excel workbook.  Scans each sheet for the first row that
    looks like Oracle column names (UPPER_CASE_WITH_UNDERSCORES), then treats
    the immediately following row as the human-readable notes row.

    Returns [{sheet_name, columns: [{column_name, position, notes}]}]
    """
    import openpyxl  # deferred import — only needed at request time

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    result = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Read first 15 rows to locate the header
        sample: list[tuple] = []
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            sample.append(row)

        header_idx: Optional[int] = None
        for idx, row in enumerate(sample):
            if _is_oracle_row(row):
                header_idx = idx
                break

        if header_idx is None:
            continue  # Skip sheets without recognisable Oracle columns

        header = sample[header_idx]
        # Notes row: the row immediately after the header, unless it also looks like a header
        notes_row: tuple = ()
        if header_idx + 1 < len(sample):
            candidate = sample[header_idx + 1]
            if not _is_oracle_row(candidate):
                notes_row = candidate

        columns = []
        for pos, col_val in enumerate(header, start=1):
            if col_val is None:
                continue
            col_name = str(col_val).strip()
            if not col_name or not _ORACLE_RE.match(col_name):
                continue
            note: Optional[str] = None
            if notes_row and pos - 1 < len(notes_row) and notes_row[pos - 1] is not None:
                raw = str(notes_row[pos - 1]).strip()
                note = raw or None
            columns.append({"column_name": col_name, "position": pos, "notes": note})

        if columns:
            result.append({"sheet_name": sheet_name, "columns": columns})

    wb.close()
    return result


# ── Pydantic models ───────────────────────────────────────────────────────────

class ColumnOut(BaseModel):
    id: uuid.UUID
    sheet_name: str
    column_name: str
    position: int
    notes: Optional[str]
    created_at: datetime

    model_config = {"from_attributes": True}


class TemplateOut(BaseModel):
    id: uuid.UUID
    name: str
    original_filename: str
    sheet_count: int
    column_count: int
    status: str
    created_at: datetime
    updated_at: datetime

    model_config = {"from_attributes": True}


class TemplateDetailOut(TemplateOut):
    columns: list[ColumnOut] = []


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/upload", response_model=TemplateDetailOut, status_code=status.HTTP_201_CREATED)
async def upload_fbdi(
    name: str = Form(...),
    file: UploadFile = File(...),
    _user_id: str = Depends(require_admin),
):
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="Only .xlsx and .xls files are supported")

    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File size must be ≤ 20 MB")

    try:
        sheets = _parse_workbook(contents)
    except Exception as exc:
        log.error("FBDI parse error for %r: %s", filename, exc)
        raise HTTPException(status_code=422, detail=f"Could not parse Excel file: {exc}")

    if not sheets:
        raise HTTPException(
            status_code=422,
            detail=(
                "No recognisable FBDI sheets found. "
                "Ensure at least one sheet has a row where column headers are UPPER_CASE_WITH_UNDERSCORES."
            ),
        )

    sheet_count = len(sheets)
    column_count = sum(len(s["columns"]) for s in sheets)

    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(_DDL)

        async with conn.transaction():
            try:
                tpl = await conn.fetchrow(
                    """
                    INSERT INTO fbdi_templates (name, original_filename, sheet_count, column_count)
                    VALUES ($1, $2, $3, $4)
                    RETURNING id, name, original_filename, sheet_count, column_count,
                              status, created_at, updated_at
                    """,
                    name.strip(), filename, sheet_count, column_count,
                )

                rows_to_insert = [
                    (tpl["id"], s["sheet_name"], c["column_name"], c["position"], c["notes"])
                    for s in sheets
                    for c in s["columns"]
                ]
                await conn.executemany(
                    """
                    INSERT INTO fbdi_columns (template_id, sheet_name, column_name, position, notes)
                    VALUES ($1, $2, $3, $4, $5)
                    """,
                    rows_to_insert,
                )

                col_rows = await conn.fetch(
                    """
                    SELECT id, sheet_name, column_name, position, notes, created_at
                    FROM fbdi_columns WHERE template_id = $1
                    ORDER BY sheet_name, position
                    """,
                    tpl["id"],
                )

            except asyncpg.UniqueViolationError:
                raise HTTPException(
                    status_code=409,
                    detail=f'A template named "{name}" already exists',
                )
            except asyncpg.PostgresError as exc:
                log.error("DB error saving FBDI template: %s", exc)
                raise HTTPException(status_code=500, detail="Database error — template not saved")

    return TemplateDetailOut.model_validate(
        {**dict(tpl), "columns": [dict(r) for r in col_rows]}
    )


@router.get("", response_model=list[TemplateOut])
async def list_templates(_user_id: str = Depends(require_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(_DDL)
        rows = await conn.fetch(
            """
            SELECT id, name, original_filename, sheet_count, column_count,
                   status, created_at, updated_at
            FROM fbdi_templates
            ORDER BY created_at DESC
            """
        )
    return [TemplateOut.model_validate(dict(r)) for r in rows]


@router.get("/{template_id}", response_model=TemplateDetailOut)
async def get_template(template_id: uuid.UUID, _user_id: str = Depends(require_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        tpl = await conn.fetchrow(
            """
            SELECT id, name, original_filename, sheet_count, column_count,
                   status, created_at, updated_at
            FROM fbdi_templates WHERE id = $1
            """,
            template_id,
        )
        if not tpl:
            raise HTTPException(status_code=404, detail="Template not found")

        col_rows = await conn.fetch(
            """
            SELECT id, sheet_name, column_name, position, notes, created_at
            FROM fbdi_columns WHERE template_id = $1
            ORDER BY sheet_name, position
            """,
            template_id,
        )

    return TemplateDetailOut.model_validate(
        {**dict(tpl), "columns": [dict(r) for r in col_rows]}
    )


@router.delete("/{template_id}")
async def delete_template(template_id: uuid.UUID, _user_id: str = Depends(require_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "DELETE FROM fbdi_templates WHERE id = $1 RETURNING id", template_id
        )
    if not row:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"deleted": True}
